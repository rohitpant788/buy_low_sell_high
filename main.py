import openpyxl
import pandas as pd
from flask import Flask, render_template, redirect, url_for

from config import WORKBOOK_FOR_BUY_LOW, WORKBOOK_V20_SHEET, WORKBOOK_V40_SHEET, WORKBOOK_ETF_SHEET
import get_nse_data

app = Flask(__name__)


@app.route("/")
def index():
    # Load data from the Excel file
    column_headers, data = load_excel_data(WORKBOOK_V20_SHEET)
    return render_template("index.html", column_headers=column_headers, data=data)


@app.route("/v20")
def v20():
    # Load data from the Excel file
    column_headers, data = load_excel_data(WORKBOOK_V20_SHEET)
    return render_template("index.html", column_headers=column_headers, data=data)


@app.route("/v40")
def v40():
    # Load data from the Excel file
    column_headers, data = load_excel_data(WORKBOOK_V40_SHEET)
    return render_template("index.html", column_headers=column_headers, data=data)


@app.route("/etf")
def etf():
    # Load data from the Excel file
    column_headers, data = load_excel_data(WORKBOOK_ETF_SHEET)
    return render_template("index.html", column_headers=column_headers, data=data)


@app.route("/reload_data", methods=["POST"])
def reload_data():
    # Call the update_workbook function for different sheets as needed
    update_workbook(WORKBOOK_V20_SHEET)
    update_workbook(WORKBOOK_V40_SHEET)
    update_workbook(WORKBOOK_ETF_SHEET)
    return redirect(url_for("index"))


def load_excel_data(sheet_name):
    # Load the workbook for buy low sell high
    wb = openpyxl.load_workbook(WORKBOOK_FOR_BUY_LOW)
    try:
        # Load the sheet with the specified name from Workbook
        sh1 = wb[sheet_name]

        # Get the total number of rows and columns
        max_row = sh1.max_row
        max_col = sh1.max_column

        # Create lists to store column headers and data
        column_headers = [sh1.cell(1, col).value for col in range(1, max_col + 1)]
        data = []

        for row in range(2, max_row + 1):  # Start from row 2 (skipping header)
            row_data = [sh1.cell(row, col).value for col in range(1, max_col + 1)]
            data.append(row_data)

        return column_headers, data
    except KeyError:
        print(f"Sheet {sheet_name} not found in the workbook.")


def update_workbook(sheet_name):
    # Load the workbook for buy low sell high
    wb = openpyxl.load_workbook(WORKBOOK_FOR_BUY_LOW)

    try:
        # Load the sheet with the specified name from Workbook
        sh1 = wb[sheet_name]

        # Get the total number of rows
        max_row = sh1.max_row

        # Create a list to store RSI and Price % from 200 DMA data and their corresponding row numbers
        rsi_data = []
        dma_data = []

        for row in range(2, max_row + 1):  # Start from row 2 (skipping header)
            # Read the symbol from the first column of each row
            symbol_cell = sh1.cell(row, 1)
            sym = symbol_cell.value

            # Add ".NS" to the symbol
            stock_symbol = sym + '.NS'

            # Get historical data
            candles = get_nse_data.get_historical_data(stock_symbol)
            df = pd.DataFrame(candles)

            print(f'Processing Row {row} for symbol {sym}')

            if not df.empty:
                # Get the last row of the DataFrame
                last_row = df.iloc[-1]

                # Update cells in the current row with relevant data
                cell_values = [
                    last_row['Close'],
                    last_row['% change for that Day'],
                    last_row['200 DMA (close)'],
                    last_row['Price % from 200 DMA'],
                    last_row['50 DMA (close)'],
                    last_row['Price < 50DMA < 200DMA'],
                    last_row['RSI']
                ]

                for col, value in enumerate(cell_values, start=2):
                    sh1.cell(row, col).value = value

                # Store RSI and Price % from 200 DMA values and row number in the lists
                rsi_data.append((last_row['RSI'], row))
                dma_data.append((last_row['Price % from 200 DMA'], row))

        # Rank the stocks based on RSI (ascending order)
        rsi_data.sort(key=lambda x: x[0])
        for rank, (rsi, row_num) in enumerate(rsi_data, start=1):
            sh1.cell(row_num, 9).value = rank  # Add RSI Rank in a new column (column I)

        # Rank the stocks based on Price % from 200 DMA (ascending order)
        dma_data.sort(key=lambda x: x[0])
        for rank, (dma, row_num) in enumerate(dma_data, start=1):
            sh1.cell(row_num, 10).value = rank  # Add 200 DMA Rank in a new column (column J)

        # Save the updated workbook
        wb.save(WORKBOOK_FOR_BUY_LOW)
    except KeyError:
        print(f"Sheet {sheet_name} not found in the workbook.")


if __name__ == "__main__":
    app.run(debug=True)
