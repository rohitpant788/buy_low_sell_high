import openpyxl
import pandas as pd
import logging

from config import WORKBOOK_FOR_BUY_LOW
import get_nse_data

# Configure logging
from logging_utils import update_log_messages

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def update_workbook(wb, sheet_name):
    try:
        sh1 = wb[sheet_name]

        # Get the total number of rows
        max_row = sh1.max_row

        # Create a list to store RSI and Price % from 200 DMA data and their corresponding row numbers
        rsi_data = []
        dma_data = []

        for row in range(2, max_row + 1):  # Start from row 2 (skipping header)
            # Read the symbol from the first column of each row
            symbol_cell = sh1.cell(row, 1)
            sym = symbol_cell.value

            # Add ".NS" to the symbol
            stock_symbol = sym + '.NS'

            # Get historical data
            candles = get_nse_data.get_historical_data(stock_symbol)
            df = pd.DataFrame(candles)

            log_message = f'Processing Row {row} for symbol {sym}'
            logger.info(log_message)
            update_log_messages(log_message)

            if not df.empty:
                # Get the last row of the DataFrame
                last_row = df.iloc[-1]

                # Update cells in the current row with relevant data
                cell_values = [
                    last_row['Close'],
                    last_row['% change for that Day'],
                    last_row['200 DMA (close)'],
                    last_row['Price % from 200 DMA'],
                    last_row['50 DMA (close)'],
                    last_row['Price < 50DMA < 200DMA'],
                    last_row['RSI']
                ]

                for col, value in enumerate(cell_values, start=2):
                    sh1.cell(row, col).value = value

                # Store RSI and Price % from 200 DMA values and row number in the lists
                rsi_data.append((last_row['RSI'], row))
                dma_data.append((last_row['Price % from 200 DMA'], row))

        # Rank the stocks based on RSI (ascending order)
        rsi_data.sort(key=lambda x: x[0])
        for rank, (rsi, row_num) in enumerate(rsi_data, start=1):
            sh1.cell(row_num, 9).value = rank  # Add RSI Rank in a new column (column I)

        # Rank the stocks based on Price % from 200 DMA (ascending order)
        dma_data.sort(key=lambda x: x[0])
        for rank, (dma, row_num) in enumerate(dma_data, start=1):
            sh1.cell(row_num, 10).value = rank  # Add 200 DMA Rank in a new column (column J)

        # Save the updated workbook
        wb.save(WORKBOOK_FOR_BUY_LOW)

    except KeyError:
        logger.error(f"Sheet {sheet_name} not found in the workbook.")
